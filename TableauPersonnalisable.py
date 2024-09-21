import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

class CustomizableInvoiceManager:
    def __init__(self):
        self.app = ctk.CTk()
        self.app.geometry("1300x620")
        self.app.title("Tableau personnalisable")

        self.products = []
        self.selected_item = None
        self.fields = []
        self.entries = {}

        self.load_form_config()
        self.setup_ui()

    def load_form_config(self):
        try:
            with open('form_config.json', 'r') as f:
                self.fields = json.load(f)
        except FileNotFoundError:
            self.fields = [
                {"name": "reference", "label": "Référence du produit", "type": "string"},
                {"name": "quantity", "label": "Quantité", "type": "number"},
                {"name": "details", "label": "Détails", "type": "string"},
                {"name": "unit_price", "label": "Prix Unitaire", "type": "number"}
            ]

    def save_form_config(self):
        with open('form_config.json', 'w') as f:
            json.dump(self.fields, f)

    def setup_ui(self):
        self.main_frame = ctk.CTkFrame(self.app, corner_radius=10)
        self.main_frame.pack(pady=20, padx=20, fill="both", expand=True)

        self.setup_input_frame()
        self.setup_table_frame()
        self.setup_action_frame()

    def setup_input_frame(self):
        if hasattr(self, 'input_frame'):
            self.input_frame.destroy()

        self.input_frame = ctk.CTkFrame(self.main_frame, corner_radius=10)
        self.input_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        label_title = ctk.CTkLabel(self.input_frame, text="Ajouter une ligne", font=ctk.CTkFont(size=16, weight="bold"))
        label_title.grid(row=0, column=0, columnspan=2, pady=10)

        self.entries.clear()
        for i, field in enumerate(self.fields):
            label = ctk.CTkLabel(self.input_frame, text=field['label'])
            label.grid(row=i+1, column=0, padx=10, pady=5, sticky="w")
            entry = ctk.CTkEntry(self.input_frame, width=200)
            entry.grid(row=i+1, column=1, padx=10, pady=5)
            self.entries[field['name']] = entry

        btn_add = ctk.CTkButton(self.input_frame, text="Ajouter la ligne", command=self.add_product)
        btn_add.grid(row=len(self.fields)+1, column=0, columnspan=2, pady=10)

    def setup_table_frame(self):
        if hasattr(self, 'table_frame'):
            self.table_frame.destroy()

        self.table_frame = ctk.CTkFrame(self.main_frame, corner_radius=10)
        self.table_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

        label_table = ctk.CTkLabel(self.table_frame, text="Tableau", font=ctk.CTkFont(size=16, weight="bold"))
        label_table.pack(pady=10)

        columns = tuple(field['label'] for field in self.fields)
        self.tree = ttk.Treeview(self.table_frame, columns=columns, show='headings', height=8)
        self.tree.pack(expand=True, fill="both", padx=10, pady=10)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, stretch=True)

        self.tree.bind("<ButtonRelease-1>", self.select_item)
        
        # Ajout d'une barre de défilement horizontale
        xscrollbar = ttk.Scrollbar(self.table_frame, orient='horizontal', command=self.tree.xview)
        xscrollbar.pack(side='bottom', fill='x')
        self.tree.configure(xscrollcommand=xscrollbar.set)

    def setup_action_frame(self):
        self.action_frame = ctk.CTkFrame(self.main_frame, corner_radius=10)
        self.action_frame.grid(row=1, column=0, columnspan=2, padx=20, pady=20, sticky="ew")

        buttons = [
            ("Modifier la ligne", self.update_product),
            ("Supprimer la ligne", self.delete_product),
            ("Exporter vers Excel", self.export_excel),
            ("Importer depuis Excel", self.import_excel),
            ("Effacer les champs", self.clear_entries),
            ("Personnaliser le formulaire", self.customize_form)
        ]

        for i, (text, command) in enumerate(buttons):
            btn = ctk.CTkButton(self.action_frame, text=text, command=command)
            btn.grid(row=0, column=i, padx=10, pady=10)

    def add_product(self):
        values = [self.entries[field['name']].get() for field in self.fields]
        if all(values):
            self.products.append(values)
            self.update_table()
            self.clear_entries()
        else:
            messagebox.showwarning("Attention", "Veuillez remplir tous les champs.")

    def update_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for product in self.products:
            self.tree.insert("", "end", values=product)

    def export_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Produits"

            # Styles
            header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            cell_font = Font(name='Arial', size=11)
            cell_alignment = Alignment(horizontal="left", vertical="center")
            
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # En-têtes
            headers = [field['label'] for field in self.fields]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Données
            for row, product in enumerate(self.products, start=2):
                for col, value in enumerate(product, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = cell_font
                    cell.alignment = cell_alignment
                    cell.border = border

                    # Appliquer un fond de couleur alterné pour les lignes
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")

            # Ajuster la largeur des colonnes
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20

            # Figer la première ligne
            ws.freeze_panes = "A2"

            wb.save(file_path)
            messagebox.showinfo("Succès", "Les données ont été exportées avec succès dans un fichier Excel stylé.")

    def import_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Récupérer les en-têtes
            headers = [cell.value for cell in ws[1]]
            
            # Mettre à jour les champs
            self.fields = [{"name": f"field_{i}", "label": header, "type": "string"} for i, header in enumerate(headers)]
            self.save_form_config()
            
            # Récupérer les données
            self.products = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.products.append(list(row))
            
            # Mettre à jour l'interface
            self.setup_input_frame()
            self.setup_table_frame()
            self.update_table()
            
            messagebox.showinfo("Succès", "Les données ont été importées avec succès depuis Excel et le formulaire a été mis à jour.")

    def clear_entries(self):
        for entry in self.entries.values():
            entry.delete(0, ctk.END)

    def select_item(self, event):
        selected = self.tree.selection()
        if selected:
            self.selected_item = self.tree.item(selected, 'values')
            for field, value in zip(self.fields, self.selected_item):
                self.entries[field['name']].delete(0, ctk.END)
                self.entries[field['name']].insert(0, value)

    def update_product(self):
        if self.selected_item:
            values = [self.entries[field['name']].get() for field in self.fields]
            if all(values):
                index = self.products.index(list(self.selected_item))
                self.products[index] = values
                self.update_table()
                self.clear_entries()
                self.selected_item = None
            else:
                messagebox.showwarning("Attention", "Veuillez remplir tous les champs pour modifier la ligne.")
        else:
            messagebox.showwarning("Attention", "Veuillez sélectionner une ligne à modifier.")

    def delete_product(self):
        if self.selected_item:
            self.products.remove(list(self.selected_item))
            self.update_table()
            self.clear_entries()
            self.selected_item = None
            messagebox.showinfo("Succès", "La ligne a été supprimé.")
        else:
            messagebox.showwarning("Attention", "Veuillez sélectionner une ligne à supprimer.")

    def customize_form(self):
        customize_window = ctk.CTkToplevel(self.app)
        customize_window.title("Personnaliser le formulaire")
        customize_window.geometry("400x400")
        customize_window.transient(self.app)  # Set the main window as the parent
        customize_window.grab_set()  # Make the window modal

        fields_frame = ctk.CTkScrollableFrame(customize_window)
        fields_frame.pack(pady=20, padx=20, fill="both", expand=True)

        self.field_entries = []
        self.field_type_vars = []  # Pour stocker les variables de type de champ
        for field in self.fields:
            frame = ctk.CTkFrame(fields_frame)
            frame.pack(fill="x", pady=5)
            
            entry = ctk.CTkEntry(frame, width=150)
            entry.insert(0, field['label'])
            entry.pack(side="left", padx=5)
            
            type_var = ctk.StringVar(value=field['type'])
            type_menu = ctk.CTkOptionMenu(frame, values=["string", "number"], variable=type_var, width=70)
            type_menu.pack(side="left", padx=5)
            
            remove_btn = ctk.CTkButton(frame, text="X", width=30, command=lambda f=frame, e=entry, t=type_var: self.remove_field(f, e, t))
            remove_btn.pack(side="left")
            
            self.field_entries.append(entry)
            self.field_type_vars.append(type_var)

        def add_field():
            frame = ctk.CTkFrame(fields_frame)
            frame.pack(fill="x", pady=5)
            
            new_entry = ctk.CTkEntry(frame, width=150)
            new_entry.pack(side="left", padx=5)
            
            new_type_var = ctk.StringVar(value="string")
            new_type_menu = ctk.CTkOptionMenu(frame, values=["string", "number"], variable=new_type_var, width=70)
            new_type_menu.pack(side="left", padx=5)
            
            remove_btn = ctk.CTkButton(frame, text="X", width=30, command=lambda f=frame, e=new_entry, t=new_type_var: self.remove_field(f, e, t))
            remove_btn.pack(side="left")
            
            self.field_entries.append(new_entry)
            self.field_type_vars.append(new_type_var)

        def remove_field(frame, entry, type_var):
            if len(self.field_entries) > 2:
                self.field_entries.remove(entry)
                self.field_type_vars.remove(type_var)
                frame.destroy()
            else:
                messagebox.showwarning("Attention", "Il doit y avoir au moins deux champs.")

        self.remove_field = remove_field

        def save_fields():
            new_fields = []
            for i, (entry, type_var) in enumerate(zip(self.field_entries, self.field_type_vars)):
                label = entry.get()
                if label:
                    new_fields.append({
                        "name": f"field_{i}",
                        "label": label,
                        "type": type_var.get()
                    })
            if len(new_fields) >= 2:
                self.fields = new_fields
                self.save_form_config()
                self.setup_input_frame()
                self.setup_table_frame()
                customize_window.destroy()
                messagebox.showinfo("Succès", "Le formulaire a été personnalisé avec succès.")
            else:
                messagebox.showwarning("Attention", "Il doit y avoir au moins deux champs.")

        btn_add_field = ctk.CTkButton(customize_window, text="Ajouter une colonne", command=add_field)
        btn_add_field.pack(pady=10)

        btn_save = ctk.CTkButton(customize_window, text="Enregistrer", command=save_fields)
        btn_save.pack(pady=10)

    def run(self):
        self.app.mainloop()

    
if __name__ == "__main__":
    app = CustomizableInvoiceManager()
    app.app.mainloop()